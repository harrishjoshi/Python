#! python3

# update_produce.py - Corrects costs in produce sales spreadsheet.
import openpyxl

wb = openpyxl.load_workbook("files/produce_sales.xlsx")
sheet = wb["Sheet"]

# The produce types and their updated prices
PRICE_UPDATES = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

# Loop through the rows and update the prices.
for row_no in range(2, sheet.max_row):  # skip the first title row
    produce_name = sheet.cell(row=row_no, column=1).value

    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_no, column=2).value = PRICE_UPDATES[produce_name]

wb.save("files/updated_produce_sales.xlsx")
